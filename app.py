# app.py

from flask import Flask, render_template, request, jsonify, redirect, url_for
from config import DATABASES, DATABASE_DIR, METADATA_FILE, load_databases
import os
import json

app = Flask(__name__)

@app.route('/')
def home():
    """메인 페이지 - DB 선택 화면"""
    # 매번 최신 DB 목록 로드
    databases = load_databases()
    return render_template('home.html', databases=databases)

@app.route('/dashboard/<db_name>')
def dashboard(db_name):
    """DB별 대시보드"""
    databases = load_databases()
    if db_name not in databases:
        return "Database not found", 404
    
    db_info = databases[db_name]
    return render_template('dashboard.html', db_name=db_name, db_info=db_info)

@app.route('/delete_db/<db_name>', methods=['POST'])
def delete_db(db_name):
    """DB 삭제"""
    databases = load_databases()
    
    if db_name not in databases:
        return jsonify({'success': False, 'message': 'DB not found'}), 404
    
    try:
        # .db 파일 삭제
        db_file = databases[db_name]['file']
        if os.path.exists(db_file):
            os.remove(db_file)
        
        # metadata.json에서 제거
        if os.path.exists(METADATA_FILE):
            with open(METADATA_FILE, 'r', encoding='utf-8') as f:
                metadata = json.load(f)
            
            if db_name in metadata:
                del metadata[db_name]
            
            with open(METADATA_FILE, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'message': f'{db_name} deleted'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
@app.route('/add_db', methods=['POST'])
def add_db():
    """새 DB 추가"""
    data = request.get_json()
    
    db_key = data.get('db_key', '').strip()
    db_name = data.get('db_name', '').strip()
    db_description = data.get('db_description', '').strip()
    db_icon = data.get('db_icon', '📁').strip()
    
    if not db_key or not db_name:
        return jsonify({'success': False, 'message': 'DB key와 이름은 필수입니다.'}), 400
    
    # 파일명으로 사용 가능한 문자인지 확인
    if not db_key.replace('_', '').isalnum():
        return jsonify({'success': False, 'message': 'DB key는 영문, 숫자, _만 사용 가능합니다.'}), 400
    
    db_file = os.path.join(DATABASE_DIR, f'{db_key}.db')
    
    # 이미 존재하는지 확인
    if os.path.exists(db_file):
        return jsonify({'success': False, 'message': '이미 존재하는 DB입니다.'}), 400
    
    try:
        # 빈 SQLite DB 생성
        import sqlite3
        conn = sqlite3.connect(db_file)
        conn.close()
        
        # metadata.json 업데이트
        if os.path.exists(METADATA_FILE):
            with open(METADATA_FILE, 'r', encoding='utf-8') as f:
                metadata = json.load(f)
        else:
            metadata = {}
        
        metadata[db_key] = {
            'name': db_name,
            'description': db_description,
            'icon': db_icon
        }
        
        with open(METADATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'message': f'{db_name} DB가 생성되었습니다.'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/api/analyze_schema/<db_name>')
def analyze_schema(db_name):
    """DB 스키마 분석"""
    from utils.schema_analyzer import analyze_schema_with_llm
    
    databases = load_databases()
    if db_name not in databases:
        return jsonify({'success': False, 'message': 'DB not found'}), 404
    
    db_path = databases[db_name]['file']
    analysis = analyze_schema_with_llm(db_path)
    
    return jsonify({'success': True, 'analysis': analysis})

@app.route('/api/suggest_queries/<db_name>')
def suggest_queries(db_name):
    """추천 질문 생성"""
    from utils.schema_analyzer import suggest_queries_with_llm
    
    databases = load_databases()
    if db_name not in databases:
        return jsonify({'success': False, 'message': 'DB not found'}), 404
    
    db_path = databases[db_name]['file']
    queries = suggest_queries_with_llm(db_path)
    
    return jsonify({'success': True, 'queries': queries})

@app.route('/api/generate_sql/<db_name>', methods=['POST'])
def generate_sql(db_name):
    """자연어 → SQL 생성"""
    from utils.query_generator import generate_sql_from_question
    
    databases = load_databases()
    if db_name not in databases:
        return jsonify({'success': False, 'message': 'DB not found'}), 404
    
    data = request.get_json()
    user_question = data.get('question', '').strip()
    model_name = data.get('model', 'gemini-2.0-flash-lite')  # 모델 선택 추가
    
    if not user_question:
        return jsonify({'success': False, 'message': '질문을 입력해주세요.'}), 400
    
    db_path = databases[db_name]['file']
    result = generate_sql_from_question(db_path, user_question, model_name)
    
    return jsonify({
        'success': True,
        'reasoning': result['reasoning'],
        'sql': result['sql']
    })

@app.route('/api/execute_sql/<db_name>', methods=['POST'])
def execute_sql_api(db_name):
    """SQL 실행"""
    from utils.query_generator import execute_sql, save_to_history
    
    databases = load_databases()
    if db_name not in databases:
        return jsonify({'success': False, 'message': 'DB not found'}), 404
    
    data = request.get_json()
    sql_query = data.get('sql', '').strip()
    question = data.get('question', '').strip()  # 질문도 함께 받기
    
    if not sql_query:
        return jsonify({'success': False, 'message': 'SQL을 입력해주세요.'}), 400
    
    db_path = databases[db_name]['file']
    result = execute_sql(db_path, sql_query)
    
    # 히스토리 저장
    if result['success']:
        save_to_history(db_name, question, sql_query, len(result['rows']))
    
    return jsonify(result)


@app.route('/api/history/<db_name>')
def get_history_api(db_name):
    """쿼리 히스토리 조회"""
    from utils.query_generator import get_history
    
    history = get_history(db_name)
    return jsonify({'success': True, 'history': history})

@app.route('/api/bookmark/<int:history_id>', methods=['POST'])
def toggle_bookmark_api(history_id):
    """북마크 토글"""
    from utils.query_generator import toggle_bookmark
    
    success = toggle_bookmark(history_id)
    return jsonify({'success': success})

@app.route('/api/export/<format>/<db_name>', methods=['POST'])
def export_data(format, db_name):
    """
    쿼리 결과를 CSV 또는 Excel로 내보내기
    
    Args:
        format: 'csv' 또는 'excel'
        db_name: DB 이름
    """
    from flask import send_file
    from utils.query_generator import execute_sql
    import csv
    import io
    from openpyxl import Workbook
    from datetime import datetime
    
    databases = load_databases()
    if db_name not in databases:
        return jsonify({'success': False, 'message': 'DB not found'}), 404
    
    data = request.get_json()
    sql_query = data.get('sql', '').strip()
    
    if not sql_query:
        return jsonify({'success': False, 'message': 'SQL을 입력해주세요.'}), 400
    
    db_path = databases[db_name]['file']
    result = execute_sql(db_path, sql_query)
    
    if not result['success']:
        return jsonify({'success': False, 'message': result['error']}), 400
    
    columns = result['columns']
    rows = result['rows']
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if format == 'csv':
        # CSV 생성
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)
        writer.writerows(rows)
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),  # UTF-8 BOM 추가 (한글 깨짐 방지)
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'{db_name}_export_{timestamp}.csv'
        )
    
    elif format == 'excel':
        # Excel 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "Query Result"
        
        # 헤더
        ws.append(columns)
        
        # 데이터
        for row in rows:
            ws.append(row)
        
        # 헤더 스타일
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # 바이트 스트림으로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{db_name}_export_{timestamp}.xlsx'
        )
    
    else:
        return jsonify({'success': False, 'message': 'Invalid format'}), 400


@app.route('/api/schema_diagram/<db_name>')
def get_schema_diagram(db_name):
    """스키마 다이어그램 (Mermaid)"""
    from utils.schema_analyzer import generate_schema_diagram
    
    databases = load_databases()
    if db_name not in databases:
        return jsonify({'success': False, 'message': 'DB not found'}), 404
    
    db_path = databases[db_name]['file']
    diagram = generate_schema_diagram(db_path)
    
    return jsonify({'success': True, 'diagram': diagram})
@app.route('/api/clear_cache/<db_name>', methods=['POST'])
def clear_schema_cache(db_name):
    """스키마 캐시 초기화"""
    from utils.schema_analyzer import clear_cache
    
    try:
        clear_cache(db_name)
        return jsonify({'success': True, 'message': f'{db_name} 캐시가 초기화되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
@app.route('/api/models')
def get_models():
    """사용 가능한 모델 목록"""
    from utils.gemini_client import get_available_models
    return jsonify({'success': True, 'models': get_available_models()})
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)